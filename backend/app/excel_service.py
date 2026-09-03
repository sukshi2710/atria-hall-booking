import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Use /tmp on Vercel/serverless environments to avoid read-only filesystem errors
if os.getenv("VERCEL"):
    EXCEL_FILE_PATH = "/tmp/college_hall_bookings_ledger.xlsx"
else:
    EXCEL_FILE_PATH = "college_hall_bookings_ledger.xlsx"


def initialize_excel_ledger():
    """Creates the formatted Excel ledger with styled headers if not already present."""
    if os.path.exists(EXCEL_FILE_PATH):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Hall Bookings Ledger"

    # Reference ID removed from headers
    headers = [
        "Faculty Name",
        "Official Mail ID",
        "Venue",
        "Department",
        "Event Details / Purpose",
        "Start Date & Time",
        "End Date & Time",
        "Status",
        "Booked At"
    ]
    
    ws.append(headers)

    # Styling header row
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Adjust column widths (shifted after removing Reference ID)
    column_widths = {
        "A": 22,  # Faculty Name
        "B": 28,  # Email
        "C": 18,  # Venue
        "D": 16,  # Department
        "E": 35,  # Event Details
        "F": 22,  # Start Time
        "G": 22,  # End Time
        "H": 15,  # Status
        "I": 22   # Booked At
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(EXCEL_FILE_PATH)


def append_booking_to_excel(booking):
    """Appends a new or updated booking record directly to the .xlsx file without the Reference ID."""
    initialize_excel_ledger()

    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    # booking_reference is recorded in the database, but omitted here from export
    row_data = [
        booking.faculty_name,
        booking.email,
        booking.venue,
        booking.department,
        booking.event_details,
        booking.start_datetime.strftime("%Y-%m-%d %I:%M %p"),
        booking.end_datetime.strftime("%Y-%m-%d %I:%M %p"),
        booking.status,
        booking.created_at.strftime("%Y-%m-%d %I:%M %p") if booking.created_at else datetime.now().strftime("%Y-%m-%d %I:%M %p")
    ]

    ws.append(row_data)

    # Status column is now column 8 (Column H)
    last_row = ws.max_row
    status_cell = ws.cell(row=last_row, column=8)
    if booking.status == "CONFIRMED":
        status_cell.font = Font(color="22543D", bold=True)
    elif booking.status == "CANCELLED":
        status_cell.font = Font(color="742A2A", bold=True)

    wb.save(EXCEL_FILE_PATH)
