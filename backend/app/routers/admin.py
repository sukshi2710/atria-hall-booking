import io
import os
from typing import List

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models import Booking, Admin
from app.schemas import BookingOut, BookingCreate
from app.auth import verify_password, create_access_token, get_current_admin
from app.conflict_engine import check_hall_clash
from app.excel_service import append_booking_to_excel
from app.email_service import send_cancellation_email

router = APIRouter(prefix="/api/v1/admin", tags=["Admin"])


@router.post("/login")
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    admin = db.query(Admin).filter(Admin.username == form_data.username).first()
    if not admin or not verify_password(form_data.password, admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token = create_access_token(data={"sub": admin.username})
    return {"access_token": access_token, "token_type": "bearer"}


@router.get("/bookings", response_model=List[BookingOut])
def list_all_bookings(admin: Admin = Depends(get_current_admin), db: Session = Depends(get_db)):
    return db.query(Booking).order_by(Booking.start_datetime.desc()).all()


@router.get("/export-excel")
def export_excel_ledger(
    admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Generates and streams the Excel ledger in-memory directly from the database."""
    bookings = db.query(Booking).order_by(Booking.start_datetime.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservations Ledger"

    headers = [
        "Venue", 
        "Faculty Name", 
        "Department", 
        "Email", 
        "Event Purpose", 
        "Start Time", 
        "End Time", 
        "Status", 
        "Created At"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for b in bookings:
        ws.append([
            b.venue,
            b.faculty_name,
            b.department,
            b.email,
            b.event_details,
            b.start_datetime.strftime("%d-%b-%Y, %I:%M %p") if b.start_datetime else "",
            b.end_datetime.strftime("%d-%b-%Y, %I:%M %p") if b.end_datetime else "",
            b.status,
            b.created_at.strftime("%d-%b-%Y, %I:%M %p") if hasattr(b, "created_at") and b.created_at else ""
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=college_hall_bookings_ledger.xlsx"
        }
    )


@router.put("/bookings/{booking_id}", response_model=BookingOut)
def update_booking(
    booking_id: int, 
    update_data: BookingCreate, 
    admin: Admin = Depends(get_current_admin), 
    db: Session = Depends(get_db)
):
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    clash = check_hall_clash(
        db, update_data.venue, update_data.start_datetime, update_data.end_datetime, exclude_booking_id=booking_id
    )
    if clash:
        raise HTTPException(status_code=409, detail=f"Cannot reschedule: Clashes with {clash.event_details}")

    for key, value in update_data.model_dump().items():
        setattr(booking, key, value)
        
    db.commit()
    db.refresh(booking)

    try:
        append_booking_to_excel(booking)
    except Exception as e:
        print(f"[EXCEL WARN] Could not sync to excel on serverless disk: {e}")

    return booking


@router.delete("/bookings/{booking_id}")
def cancel_booking(
    booking_id: int,
    admin: Admin = Depends(get_current_admin), 
    db: Session = Depends(get_db)
):
    booking = db.query(Booking).filter(Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    booking.status = "CANCELLED"
    db.commit()
    db.refresh(booking)

    try:
        append_booking_to_excel(booking)
    except Exception as e:
        print(f"[EXCEL WARN] Could not write to Excel: {e}")

    if booking.email:
        try:
            send_cancellation_email(
                to_email=booking.email,
                faculty_name=booking.faculty_name,
                venue=booking.venue,
                event_details=booking.event_details,
                start_time=booking.start_datetime.strftime("%d-%b-%Y, %I:%M %p"),
                end_time=booking.end_datetime.strftime("%d-%b-%Y, %I:%M %p")
            )
        except Exception as e:
            print(f"[EMAIL ERROR] Failed during cancellation dispatch: {e}")

    return {"message": "Booking successfully cancelled and notification email sent"}
